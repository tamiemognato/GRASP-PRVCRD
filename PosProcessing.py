import matplotlib
import openpyxl
import pandas as pd
from matplotlib import pyplot as plt


from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

from Functions.SavingsHeuristicClarkeAndWrightFunctions import calcula_ROUTES_DEMANDAS_E_TOTAL_POINTS, \
    calcula_custo_total

import pandas as pd
import matplotlib.pyplot as plt


def GerarGraficosIteracoesCustos(arquivo):
    # Ler os dados do Excel para um DataFrame
    df = pd.read_excel(arquivo)

    # Iterar sobre os valores únicos em cada coluna relevante
    for alfa in df['ALFA'].unique():
        for semente in df['SEMENTE'].unique():
            for max_iter in df['NumeroMaximoIteracoesSemMelhoria'].unique():
                # Filtrar o DataFrame para os valores específicos de alfa, semente e max_iter
                filtro = (df['ALFA'] == alfa) & (df['SEMENTE'] == semente) & (
                            df['NumeroMaximoIteracoesSemMelhoria'] == max_iter)
                dados_filtrados = df[filtro]

                # Criar um gráfico de pontos ligados por retas
                plt.figure()
                plt.plot(dados_filtrados['ITERAÇÃO'], dados_filtrados['CUSTOS'], marker='o')

                # Definir título do gráfico
                titulo = f'Iteração x Custos - ALFA {alfa} - MaxIter {max_iter} - Semente {semente}'
                plt.title(titulo)

                # Definir rótulos dos eixos
                plt.xlabel('ITERAÇÃO')
                plt.ylabel('CUSTOS')

                # Salvar o gráfico com o mesmo nome que o título
                nome_arquivo = titulo + '.png'
                plt.savefig(nome_arquivo)
                plt.close()

    return


def CriarExcel_lista_registros_melhoras_GRASP(listas_lista_registros_melhoras,file):
    # Criar um DataFrame
    lista_tuplas = []

    for lista in listas_lista_registros_melhoras:
        for tupla in lista:
            lista_tuplas.append(tupla)

    df = pd.DataFrame(lista_tuplas, columns=["ALFA","SEMENTE","NumeroMaximoIteracoesSemMelhoria","CUSTOS", "ITERAÇÃO", "improvement_time", "elapsed_time_LCR", "elapsed_time_AdTestes", "elapsed_time_savings_inteiro","elapsed_time_doisopt_inteiro"])

    # Salvar o DataFrame em um arquivo Excel
    df.to_excel(file + '_registrosMelhoras.xlsx', index=False)

    return


def GerarDicionarioSolucao(MelhorSolucao,elapsed_time,demanda,custos,nome_instancia,s,alfa,semente,maxtime,capacidade,NumeroMaximoIteracoesSemMelhoria):

    listaROUTES_DEMANDS, listaTOTAL_POINTS = calcula_ROUTES_DEMANDAS_E_TOTAL_POINTS(MelhorSolucao, demanda)
    custo_total, custo_por_rota = calcula_custo_total(custos, MelhorSolucao)
    dic_solucao = {'NAME': nome_instancia,
                  'TYPE': 'DOIS_OPT',
                  'SAFETY_STOCK': s,
                  'ALFA': alfa,
                  'SEMENTE': semente,
                  'MAXTIME': maxtime,
                  'CAPACITY': capacidade,
                  'COST': custo_total,
                  'QUANTITY_ROUTES': len(MelhorSolucao),
                  'ROUTES': MelhorSolucao,
                  'ROUTES_COSTS': custo_por_rota,
                  'ROUTES_DEMANDS': listaROUTES_DEMANDS,
                  'TOTAL_POINTS': listaTOTAL_POINTS,
                  'ELAPSED_TIME': elapsed_time,
                  'ITERAÇÕES_GRASP':NumeroMaximoIteracoesSemMelhoria}

    return dic_solucao

def CriarExcelAlfasSavings(ListaDeDicionariosDosArquivos,file):
    # Crie um novo arquivo Excel
    workbook = Workbook()

    # Selecione a planilha ativa (por padrão, é a primeira planilha)
    sheet = workbook.active

    # Defina o nome da aba como "TabelaComparativa"
    sheet.title = "AlfasSavings"

    # Defina os cabeçalhos das colunas
    headers = ["Alfa", "Custo", "Tempo", "Semente"]

    # Adicione os cabeçalhos na primeira linha da planilha
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    # Adicione os cabeçalhos na primeira linha da planilha
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header



    # Preencha as linhas com os valores dos dicionários
    for row_num, dicionario in enumerate(ListaDeDicionariosDosArquivos, 2):
        for col_num, chave in enumerate(headers, 1):
            sheet.cell(row=row_num, column=col_num).value = dicionario[chave]

    # Salve o arquivo Excel
    workbook.save(file + "AlfasSavings.xlsx")

    return

def CriarExcelTabelaComparativa_GRASPMaxTime(ListaDeDicionariosDosArquivos,file):
    # Crie um novo arquivo Excel
    workbook = Workbook()

    # Selecione a planilha ativa (por padrão, é a primeira planilha)
    sheet = workbook.active

    # Defina o nome da aba como "TabelaComparativa"
    sheet.title = "TabelaComparativa"

    # Defina os cabeçalhos das colunas
    headers = ["NAME", "TYPE", "SAFETY_STOCK", "ALFA", "SEMENTE", "MAXTIME", "CAPACITY", "COST", "QUANTITY_ROUTES", "ELAPSED_TIME", "Erro", "ITERAÇÕES_GRASP", "ContadorTempoMaximo","TempoMaximo","iteracoesMelhorSolucao","tempoMelhorSolucao"]

    # Adicione os cabeçalhos na primeira linha da planilha
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header



    # Preencha as linhas com os valores dos dicionários
    for row_num, dicionario in enumerate(ListaDeDicionariosDosArquivos, 2):
        for col_num, chave in enumerate(headers, 1):
            sheet.cell(row=row_num, column=col_num).value = dicionario[chave]

    # Salve o arquivo Excel
    workbook.save(file + "_TabelaResultadosGRASP.xlsx")

    return


def CriarExcelTabelaComparativa_GRASP(ListaDeDicionariosDosArquivos,file):
    # Crie um novo arquivo Excel
    workbook = Workbook()

    # Selecione a planilha ativa (por padrão, é a primeira planilha)
    sheet = workbook.active

    # Defina o nome da aba como "TabelaComparativa"
    sheet.title = "TabelaComparativa"

    # Defina os cabeçalhos das colunas
    headers = ["NAME", "TYPE", "SAFETY_STOCK", "ALFA", "SEMENTE", "MAXTIME", "CAPACITY", "COST", "QUANTITY_ROUTES", "ELAPSED_TIME", "Erro", "ITERAÇÕES_GRASP"]

    # Adicione os cabeçalhos na primeira linha da planilha
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header



    # Preencha as linhas com os valores dos dicionários
    for row_num, dicionario in enumerate(ListaDeDicionariosDosArquivos, 2):
        for col_num, chave in enumerate(headers, 1):
            sheet.cell(row=row_num, column=col_num).value = dicionario[chave]

    # Salve o arquivo Excel
    workbook.save(file + "_TabelaResultadosGRASP.xlsx")

    return

def CriarExcelTabelaComparativa_R(ListaDeDicionariosDosArquivos, name):
    # Crie um novo arquivo Excel
    workbook = Workbook()

    # Selecione a planilha ativa (por padrão, é a primeira planilha)
    sheet = workbook.active

    # Defina o nome da aba como "TabelaComparativa"
    sheet.title = "TabelaComparativa"

    # Defina os cabeçalhos das colunas
    headers = ["NAME", "TYPE", "SAFETY_STOCK", "ALFA", "SEMENTE", "MAXTIME", "CAPACITY", "COST", "QUANTITY_ROUTES", "elapsed_time_seg", "Erro"]

    # Adicione os cabeçalhos na primeira linha da planilha
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header



    # Preencha as linhas com os valores dos dicionários
    for row_num, dicionario in enumerate(ListaDeDicionariosDosArquivos, 2):
        for col_num, chave in enumerate(headers, 1):
            sheet.cell(row=row_num, column=col_num).value = dicionario[chave]

    # Salve o arquivo Excel
    workbook.save(name + "_TabelaResultados.xlsx")

    return


def ExcelTabelaComparativaGeral(files_inputrun):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resultados"
    headers =[ "NAME", "TYPE", "SAFETY_STOCK", "ALFA", "SEMENTE", "MAXTIME", "CAPACITY", "COST", "QUANTITY_ROUTES", "ELAPSED_TIME", "Erro","ITERAÇÕES_GRASP"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    # Abre cada excel do arquivo contido na lista files_inputrun
    for file_inputrun in files_inputrun:
        # Carrega o arquivo do excel
        wb = openpyxl.load_workbook(file_inputrun)
        ws = wb.active

        # Lê os valores das linhas do excel
        for row_num in range(2, ws.max_row + 1):
            values = []
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                values.append(cell.value)

            # Adiciona os valores das linhas no novo excel sendo gerado
            sheet.append(values)

    workbook.save("TodasInstancias_TabelaResuldadosGRASP.xlsx")


def CriarExcelTabelaComparativa(ListaDeDicionariosDosArquivos):
    # Crie um novo arquivo Excel
    workbook = Workbook()

    # Selecione a planilha ativa (por padrão, é a primeira planilha)
    sheet = workbook.active

    # Defina o nome da aba como "TabelaComparativa"
    sheet.title = "TabelaComparativa"

    # Defina os cabeçalhos das colunas
    headers = ["NAME", "TYPE", "SAFETY_STOCK", "MAXTIME", "COST", "QUANTITY_ROUTES"]

    # Adicione os cabeçalhos na primeira linha da planilha
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header


    # Preencha as linhas com os valores dos dicionários
    for row_num, dicionario in enumerate(ListaDeDicionariosDosArquivos, 2):
        for col_num, chave in enumerate(headers, 1):
            sheet.cell(row=row_num, column=col_num).value = dicionario[chave]

    # Salve o arquivo Excel
    workbook.save("TabelaResultados.xlsx")

    return



#Pensar para cada laço dos savings plotar um grafico parcial para ver como a solução esta evoluindo, isso é importante na busca local para ver como a solução esta evoluindo
def GerarGraficoRotas_EVRPST(coordenadas,rotas,nome_instancia,s):
    plt.clf()  # Limpa a figura atual

    pontos_rota = []
    for rota in rotas:
        pontos = [coordenadas[i] for i in rota]
        pontos_rota.append(pontos)

    # Desenhar o gráfico
    for pontos in pontos_rota:
        x = [p[0] for p in pontos]
        y = [p[1] for p in pontos]
        plt.plot(x, y, marker='o')

        # Adicionar setas nas linhas
        for i in range(len(x) - 1):
            dx = x[i + 1] - x[i]
            dy = y[i + 1] - y[i]
            plt.arrow(x[i], y[i], dx, dy, color='black', length_includes_head=True,head_width=1.5, head_length=1.5)

    # Adicionar os pontos individuais
    x = [p[0] for p in coordenadas]
    y = [p[1] for p in coordenadas]
    plt.scatter(x, y, color='black', zorder=10,s=150)

    for i, ponto in enumerate(coordenadas):
        plt.text(ponto[0], ponto[1], str(i), horizontalalignment='center', verticalalignment='center', zorder=11,fontweight='bold',color='white')

    plt.xlabel('Coordenada X')
    plt.ylabel('Coordenada Y')
    plt.title('Rotas EVRPST '+ nome_instancia)

    # Salvar a figura como arquivo
    plt.savefig(nome_instancia +'_'+ str(s) +'_EVRPST.png', dpi=300, bbox_inches='tight')


    #plt.show()

    return

def GerarGraficoRotas_EVRPST_r(coordenadas,rotas,nome_instancia,s,semente,alfa):
    plt.clf()  # Limpa a figura atual

    pontos_rota = []
    for rota in rotas:
        pontos = [coordenadas[i] for i in rota]
        pontos_rota.append(pontos)

    # Desenhar o gráfico
    for pontos in pontos_rota:
        x = [p[0] for p in pontos]
        y = [p[1] for p in pontos]
        plt.plot(x, y, marker='o')

        # Adicionar setas nas linhas
        for i in range(len(x) - 1):
            dx = x[i + 1] - x[i]
            dy = y[i + 1] - y[i]
            plt.arrow(x[i], y[i], dx, dy, color='black', length_includes_head=True,head_width=1.5, head_length=1.5)

    # Adicionar os pontos individuais
    x = [p[0] for p in coordenadas]
    y = [p[1] for p in coordenadas]
    plt.scatter(x, y, color='black', zorder=10,s=150)

    for i, ponto in enumerate(coordenadas):
        plt.text(ponto[0], ponto[1], str(i), horizontalalignment='center', verticalalignment='center', zorder=11,fontweight='bold',color='white')

    plt.xlabel('Coordenada X')
    plt.ylabel('Coordenada Y')
    plt.title('Rotas EVRPST '+ nome_instancia)

    # Salvar a figura como arquivo
    plt.savefig(nome_instancia +'_'+ str(s) +'_'+ str(semente)+'_'+ str(alfa)  + '_EVRPST.png', dpi=300, bbox_inches='tight')

    #plt.show()

    return


def GerarGraficoRotas_CVRP(coordenadas,rotas,nome_instancia):
    plt.clf()  # Limpa a figura atual

    pontos_rota = []
    for rota in rotas:
        pontos = [coordenadas[i] for i in rota]
        pontos_rota.append(pontos)

    # Desenhar o gráfico
    for pontos in pontos_rota:
        x = [p[0] for p in pontos]
        y = [p[1] for p in pontos]
        plt.plot(x, y, marker='o')

        # Adicionar setas nas linhas
        for i in range(len(x) - 1):
            dx = x[i + 1] - x[i]
            dy = y[i + 1] - y[i]
            plt.arrow(x[i], y[i], dx, dy, color='black', length_includes_head=True,head_width=1.5, head_length=1.5)

    # Adicionar os pontos individuais
    x = [p[0] for p in coordenadas]
    y = [p[1] for p in coordenadas]
    plt.scatter(x, y, color='black', zorder=10,s=150)

    for i, ponto in enumerate(coordenadas):
        plt.text(ponto[0], ponto[1], str(i), horizontalalignment='center', verticalalignment='center', zorder=11,fontweight='bold',color='white')

    plt.xlabel('Coordenada X')
    plt.ylabel('Coordenada Y')
    plt.title('Rotas CVRP '+ nome_instancia)

    # Salvar a figura como arquivo
    plt.savefig(nome_instancia +'_CVRP.png', dpi=300, bbox_inches='tight')


    #plt.show()

    return



def Testa_Solucao_Retorna_Erros_EVRPST(dic_instance,dic_routes):
    erro = False
    for cost in dic_routes['ROUTES_COSTS']:
        if cost > dic_routes['MAXTIME']:
            print('ERRO: Custo de rota maior que o maxtime.')
            erro = True


    for demand in dic_routes['ROUTES_DEMANDS']:
        if demand > dic_instance['CAPACITY']:
            print('ERRO: Demanda da rota maior que a capacidade.')
            erro = True


    contador = 0
    soma = 0
    # print("len(dic_routes['TOTAL_POINTS']): ",len(dic_routes['TOTAL_POINTS']))
    while contador < len(dic_routes['TOTAL_POINTS']):
        # print("contador: ",contador)
        # print("dic_routes['TOTAL_POINTS'][contador]: ", dic_routes['TOTAL_POINTS'][contador])
        soma = soma + dic_routes['TOTAL_POINTS'][contador]
        # print("soma: ",soma)
        contador += 1

    if soma != (dic_instance['DIMENSION']-1):
        # print("soma: ",soma)
        # print("(dic_instance['DIMENSION']-1): ",(dic_instance['DIMENSION']-1))
        print('ERRO: Não atendeu a todos os clientes.')
        erro = True

    return erro


def Testa_Solucao_Retorna_Erros_CVRP(dic_instance,dic_routes):
    for demand in dic_routes['ROUTES_DEMANDS']:
        if demand > dic_instance['CAPACITY']:
            print('ERRO: Demanda da rota maior que a capacidade.')

    contador = 0
    soma = 0
    while contador < len(dic_routes['TOTAL_POINTS']):
        soma = soma + dic_routes['TOTAL_POINTS'][contador]
        contador += 1

    if soma != (dic_instance['DIMENSION']-1):
        print('ERRO: Não atendeu a todos os clientes.')


    return